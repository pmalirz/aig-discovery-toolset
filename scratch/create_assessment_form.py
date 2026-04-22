import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import RadarChart, Reference
from openpyxl.utils import get_column_letter

def create_excel():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Assessment Form"

    # Styles
    title_font = Font(bold=True, size=16, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="002060")
    dim_font = Font(bold=True, size=12)
    dim_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="BFBFBF")
    bold_font = Font(bold=True)
    subtotal_fill = PatternFill("solid", fgColor="F2F2F2")
    avg_fill = PatternFill("solid", fgColor="E2EFDA")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Column widths
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 60

    # Title
    sheet.merge_cells('A1:F1')
    sheet['A1'] = "AI Opportunity Assessment Form"
    sheet['A1'].font = title_font
    sheet['A1'].fill = title_fill
    sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30

    # Basic Info
    for r, label in [(3, "Idea / Project Name:"), (4, "Submitted By (Team):"), (5, "Date:")]:
        sheet[f'B{r}'] = label
        sheet[f'B{r}'].font = header_font
        sheet[f'C{r}'].border = thin_border
        sheet.merge_cells(f'C{r}:F{r}')

    # Table Headers
    row = 7
    headers = ["#", "Criterion", "Weight", "Score (1-5)", "Score x Weight", "Evidence / Notes"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Dimension definitions
    dimensions = [
        {
            "name": "1. Business Value & Strategic Fit",
            "weight": 0.30,
            "criteria": [
                ("1a. Magnitude of business impact", 0.35),
                ("1b. Strategic alignment", 0.15),
                ("1c. Scalability of value", 0.10),
                ("1d. Time-to-value", 0.15),
                ("1e. Competitive differentiation", 0.10),
                ("1f. End-User Desirability & Workflow Fit", 0.15)
            ]
        },
        {
            "name": "2. Data Readiness & Availability",
            "weight": 0.20,
            "criteria": [
                ("2a. Data existence & accessibility", 0.30),
                ("2b. Data volume & representativeness", 0.25),
                ("2c. Data quality", 0.25),
                ("2d. Data governance & ownership", 0.20)
            ]
        },
        {
            "name": "3. Technical Feasibility & Architecture Fit",
            "weight": 0.15,
            "criteria": [
                ("3a. Solution maturity", 0.20),
                ("3b. Integration complexity", 0.20),
                ("3c. Build vs buy decision clarity", 0.15),
                ("3d. Infrastructure & MLOps readiness", 0.25),
                ("3e. Vendor / technology lock-in risk", 0.20)
            ]
        },
        {
            "name": "4. Cost, Effort & ROI",
            "weight": 0.15,
            "criteria": [
                ("4a. Implementation cost (CAPEX)", 0.25),
                ("4b. Ongoing operational cost (OPEX)", 0.20),
                ("4c. Estimated ROI & payback period", 0.35),
                ("4d. Organisational effort (change burden)", 0.20)
            ]
        },
        {
            "name": "5. Risk & Ethical Profile",
            "weight": 0.10,
            "criteria": [
                ("5a. Regulatory & compliance risk", 0.20),
                ("5b. AI output quality risk", 0.25),
                ("5c. Data privacy & security risk", 0.20),
                ("5d. Adoption & change resistance risk", 0.15),
                ("5e. Dependency & concentration risk", 0.05),
                ("5f. Explainability, Fairness & Ethical Risk", 0.15)
            ]
        },
        {
            "name": "6. Organisational & Change Readiness",
            "weight": 0.10,
            "criteria": [
                ("6a. Executive sponsorship", 0.30),
                ("6b. Internal AI capability & talent", 0.25),
                ("6c. Digital & data culture maturity", 0.20),
                ("6d. AI governance framework", 0.25)
            ]
        }
    ]

    # Data Validation for Scores (1-5)
    dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    sheet.add_data_validation(dv)

    row += 1
    dimension_subtotal_cells = []
    dimension_avg_cells = []       # For the radar chart
    dimension_short_names = []     # Labels for the radar chart

    for dim in dimensions:
        num_criteria = len(dim['criteria'])

        # Dimension Header
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        dim_cell = sheet.cell(row=row, column=1)
        dim_cell.value = f"{dim['name']} ({dim['weight']*100:.0f}%)"
        dim_cell.font = dim_font
        dim_cell.fill = dim_fill
        for col_num in range(1, 7):
            c = sheet.cell(row=row, column=col_num)
            c.border = thin_border
            if col_num > 2:
                c.fill = dim_fill

        dim_weight_cell = sheet.cell(row=row, column=3)
        dim_weight_cell.value = dim['weight']
        dim_weight_cell.number_format = '0%'
        dim_weight_cell.font = dim_font
        dim_weight_cell.alignment = Alignment(horizontal="center")

        dim_header_row = row
        row += 1
        start_crit_row = row

        for crit_name, crit_weight in dim['criteria']:
            sheet.cell(row=row, column=2).value = crit_name
            weight_cell = sheet.cell(row=row, column=3)
            weight_cell.value = crit_weight
            weight_cell.number_format = '0%'
            weight_cell.alignment = Alignment(horizontal="center")

            score_cell = sheet.cell(row=row, column=4)
            score_cell.alignment = Alignment(horizontal="center")
            score_cell.font = Font(color="0000FF", bold=True)
            dv.add(score_cell)

            calc_cell = sheet.cell(row=row, column=5)
            calc_cell.value = f"=C{row}*D{row}"
            calc_cell.number_format = '0.00'
            calc_cell.alignment = Alignment(horizontal="center")

            for col_num in range(1, 7):
                sheet.cell(row=row, column=col_num).border = thin_border
            row += 1

        end_crit_row = row - 1

        # Weighted Subtotal row
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        st_cell = sheet.cell(row=row, column=1)
        st_cell.value = f"Weighted Subtotal"
        st_cell.font = Font(italic=True, bold=True)
        st_cell.alignment = Alignment(horizontal="right")
        st_score = sheet.cell(row=row, column=5)
        st_score.value = f"=SUM(E{start_crit_row}:E{end_crit_row})*C{dim_header_row}"
        st_score.number_format = '0.00'
        st_score.font = bold_font
        st_score.alignment = Alignment(horizontal="center")
        dimension_subtotal_cells.append(f"E{row}")
        for col_num in range(1, 7):
            sheet.cell(row=row, column=col_num).border = thin_border
            sheet.cell(row=row, column=col_num).fill = subtotal_fill
        row += 1

        # Average Score row (raw / 5 — for radar chart)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        avg_label = sheet.cell(row=row, column=1)
        avg_label.value = f"Avg Score (out of 5)"
        avg_label.font = Font(italic=True, bold=True, color="006100")
        avg_label.alignment = Alignment(horizontal="right")

        avg_cell = sheet.cell(row=row, column=4)
        avg_cell.value = f"=IFERROR(SUM(D{start_crit_row}:D{end_crit_row})/{num_criteria},0)"
        avg_cell.number_format = '0.00'
        avg_cell.font = Font(bold=True, color="006100")
        avg_cell.alignment = Alignment(horizontal="center")
        dimension_avg_cells.append(f"D{row}")

        # Extract short name for radar label (remove number prefix)
        short = dim['name'].split('. ', 1)[1] if '. ' in dim['name'] else dim['name']
        dimension_short_names.append(short)

        for col_num in range(1, 7):
            sheet.cell(row=row, column=col_num).border = thin_border
            sheet.cell(row=row, column=col_num).fill = avg_fill

        row += 2  # gap

    # ─── TOTAL SCORE ───
    total_row = row
    sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    tot_label = sheet.cell(row=total_row, column=1)
    tot_label.value = "TOTAL SCORE (Out of 5.00):"
    tot_label.font = Font(bold=True, size=14)
    tot_label.alignment = Alignment(horizontal="right")

    tot_val = sheet.cell(row=total_row, column=5)
    tot_val.value = f"={'+'.join(dimension_subtotal_cells)}"
    tot_val.font = Font(bold=True, size=14, color="008000")
    tot_val.number_format = '0.00'
    tot_val.alignment = Alignment(horizontal="center")

    for col_num in range(1, 6):
        c = sheet.cell(row=total_row, column=col_num)
        c.border = thin_border
        c.fill = PatternFill("solid", fgColor="D9E1F2")

    row += 2

    # ─── VERDICT ───
    verdict_row = row
    sheet.merge_cells(start_row=verdict_row, start_column=1, end_row=verdict_row, end_column=4)
    sheet.cell(row=verdict_row, column=1).value = "Verdict:"
    sheet.cell(row=verdict_row, column=1).font = Font(bold=True, size=12)
    sheet.cell(row=verdict_row, column=1).alignment = Alignment(horizontal="right")

    tr = total_row
    verd_val = sheet.cell(row=verdict_row, column=5)
    verd_val.value = f'=IF(E{tr}>=4.25,"\u2705 Proceed",IF(E{tr}>=3.5,"\U0001f7e2 Proceed with conditions",IF(E{tr}>=2.75,"\U0001f7e1 Pilot / de-risk first",IF(E{tr}>=2,"\U0001f7e0 Defer","\U0001f534 Do not proceed"))))'
    verd_val.font = Font(bold=True, size=12)
    verd_val.alignment = Alignment(horizontal="center")
    for col_num in range(1, 6):
        sheet.cell(row=verdict_row, column=col_num).border = thin_border

    row += 1

    # ─── ACTION ───
    action_row = row
    sheet.merge_cells(start_row=action_row, start_column=1, end_row=action_row, end_column=4)
    sheet.cell(row=action_row, column=1).value = "Action:"
    sheet.cell(row=action_row, column=1).font = Font(bold=True, size=12)
    sheet.cell(row=action_row, column=1).alignment = Alignment(horizontal="right")

    act_val = sheet.cell(row=action_row, column=5)
    act_val.value = f'=IF(E{tr}>=4.25,"Strong case. Prioritise and move to solution design.",IF(E{tr}>=3.5,"Good potential. Address identified gaps first.",IF(E{tr}>=2.75,"Mixed picture. Run a focused proof of concept.",IF(E{tr}>=2,"Significant blockers. Create a remediation roadmap.","Fundamental blockers. Revisit in 12 months."))))'
    act_val.alignment = Alignment(horizontal="center", wrap_text=True)
    sheet.row_dimensions[action_row].height = 45
    for col_num in range(1, 6):
        sheet.cell(row=action_row, column=col_num).border = thin_border

    # ─── CONDITIONAL FORMATTING ───
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    light_green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    orange_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_font = Font(color='006100', bold=True)
    yellow_font = Font(color='9C5700', bold=True)
    orange_font = Font(color='C65911', bold=True)
    red_font = Font(color='9C0006', bold=True)

    sheet.conditional_formatting.add(f"E{verdict_row}", CellIsRule(operator='equal', formula=['"\u2705 Proceed"'], stopIfTrue=True, fill=green_fill, font=green_font))
    sheet.conditional_formatting.add(f"E{verdict_row}", CellIsRule(operator='equal', formula=['"\U0001f7e2 Proceed with conditions"'], stopIfTrue=True, fill=light_green_fill, font=green_font))
    sheet.conditional_formatting.add(f"E{verdict_row}", CellIsRule(operator='equal', formula=['"\U0001f7e1 Pilot / de-risk first"'], stopIfTrue=True, fill=yellow_fill, font=yellow_font))
    sheet.conditional_formatting.add(f"E{verdict_row}", CellIsRule(operator='equal', formula=['"\U0001f7e0 Defer"'], stopIfTrue=True, fill=orange_fill, font=orange_font))
    sheet.conditional_formatting.add(f"E{verdict_row}", CellIsRule(operator='equal', formula=['"\U0001f534 Do not proceed"'], stopIfTrue=True, fill=red_fill, font=red_font))

    # ─── RADAR CHART DATA TABLE (hidden helper area in column H-I) ───
    radar_label_col = 8   # H
    radar_value_col = 9   # I
    radar_start_row = 2

    sheet.cell(row=radar_start_row, column=radar_label_col).value = "Dimension"
    sheet.cell(row=radar_start_row, column=radar_label_col).font = bold_font
    sheet.cell(row=radar_start_row, column=radar_value_col).value = "Avg Score"
    sheet.cell(row=radar_start_row, column=radar_value_col).font = bold_font

    for i, (name, avg_ref) in enumerate(zip(dimension_short_names, dimension_avg_cells)):
        r = radar_start_row + 1 + i
        sheet.cell(row=r, column=radar_label_col).value = name
        sheet.cell(row=r, column=radar_value_col).value = f"={avg_ref}"
        sheet.cell(row=r, column=radar_value_col).number_format = '0.00'

    radar_end_row = radar_start_row + len(dimension_short_names)

    sheet.column_dimensions['H'].width = 35
    sheet.column_dimensions['I'].width = 12

    # ─── RADAR CHART ───
    chart = RadarChart()
    chart.type = "filled"
    chart.title = "Dimension Scores (Avg out of 5)"
    chart.style = 1
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 5

    labels = Reference(sheet, min_col=radar_label_col, min_row=radar_start_row + 1, max_row=radar_end_row)
    data = Reference(sheet, min_col=radar_value_col, min_row=radar_start_row, max_row=radar_end_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    chart.width = 20
    chart.height = 16

    # Place chart to the right of the summary table, aligned with total score row
    sheet.add_chart(chart, f"H{total_row}")

    # Save
    output_path = os.path.join(os.getcwd(), 'tracker', 'templates', 'AI_Opportunity_Assessment_Form.xlsx')
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Created Excel form at {output_path}")

if __name__ == '__main__':
    create_excel()
